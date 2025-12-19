import pandas as pd
from openpyxl.utils import range_boundaries, get_column_letter


def _parse_range(cell_range: str):
    """
    Convert an Excel-style range string (e.g., "A1:C5") into parameters for pandas:
      - usecols: column range like "A:C"
      - skiprows: number of rows to skip before reading
      - nrows:    number of rows to read

    Args:
        cell_range (str): Excel range, e.g. "B2:D10".
    Returns:
        tuple: (usecols:str, skiprows:int, nrows:int)
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
    skiprows = min_row - 1
    nrows = max_row - min_row + 1
    return usecols, skiprows, nrows


def read_1d_dict(
    filepath: str,
    sheet_name: str,
    cell_range: str
) -> dict:
    """
    Read a 1D Excel range into a Python dict.

    This reads values in row-major order (flattened). Keys are integer indices starting at 0.

    Example:
        A2:A11 (10 rows) -> {0: val_at_A2, 1: val_at_A3, ..., 9: val_at_A11}
    """
    usecols, skiprows, nrows = _parse_range(cell_range)
    df = pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=None,
        usecols=usecols,
        skiprows=skiprows,
        nrows=nrows
    )
    flat = df.values.flatten()
    return {i: flat[i] for i in range(len(flat))}


def read_2d_dict(
    filepath: str,
    sheet_name: str,
    cell_range: str
) -> dict:
    """
    Read a 2D Excel table into a Python dict.

    Keys are tuples (row_idx, col_idx), zero-based.

    Example:
        B2:D4 (3 rows x 3 cols) -> keys from (0,0) to (2,2)
    """
    usecols, skiprows, nrows = _parse_range(cell_range)
    df = pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=None,
        usecols=usecols,
        skiprows=skiprows,
        nrows=nrows
    )
    arr = df.values
    return {
        (i, j): arr[i, j]
        for i in range(arr.shape[0])
        for j in range(arr.shape[1])
    }


def read_3d_dict(
    filepath: str,
    sheet_name: str,
    cell_range: str,
    dim_sizes: tuple
) -> dict:
    """
    Read a 3D Excel range into a Python dict.

    Excel range must cover exactly d1*d2*d3 cells in row-major order:
      first varying fastest along the last dimension.

    Keys are tuples (i, j, k) with 0 <= i < d1, 0 <= j < d2, 0 <= k < d3.
    Values are assigned from the flattened list.

    Example order for dim_sizes=(d1,d2,d3):
      flat index = i*(d2*d3) + j*(d3) + k
    """
    usecols, skiprows, nrows = _parse_range(cell_range)
    df = pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=None,
        usecols=usecols,
        skiprows=skiprows,
        nrows=nrows
    )
    flat = df.values.flatten()
    d1, d2, d3 = dim_sizes
    expected = d1 * d2 * d3
    if len(flat) != expected:
        raise ValueError(
            f"Expected {expected} values for dimensions {dim_sizes}, but got {len(flat)}."
        )
    return {
        (i, j, k): flat[i * d2 * d3 + j * d3 + k]
        for i in range(d1)
        for j in range(d2)
        for k in range(d3)
    }
